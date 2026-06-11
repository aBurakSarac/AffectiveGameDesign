"""Annotation clustering algorithms for the La Façade Fissuréе FER pipeline.

Pattern: [Strategy] — three pluggable annotation algorithms:
    v1  peak detection + fixed 6s window (original; predictable rhythm)
    v2  flood-fill clustering with confidence scoring (adaptive width)
    v3  multi-channel voting with 8 independent signal channels
"""

import csv
import sys
from pathlib import Path

from fer.annotation_events import (
    SEED_EMOTIONS, HS_EMOTIONS, RPPG_ALGOS,
    HEADER_V1, HEADER_V2, HEADER_V3,
    CHANNEL_SOURCE, CHANNEL_EMOTION_VOTE, FORMULA_VOTE_THR,
    CATEGORY_OPTIONS,
    _count_emotions, _all_formula_scores, _channels_active_at,
)
from fer.annotation_io import (
    _output_dir, load_scene_cuts, load_rppg_csv, rppg_cluster_stats,
)


def _make_fn_rows(header, n):
    """Build n blank FN placeholder rows for a given header list."""
    rows = []
    for j in range(1, n + 1):
        row = {col: "" for col in header}
        row["event_id"] = f"FN_{j:02d}"
        rows.append(row)
    return rows


def _write_csv(results, header, out_path):
    """Write annotation results to CSV."""
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        writer.writerows(results)


# ══════════════════════════════════════════════════════════════════════════
# v1: Peak detection + fixed window
# ══════════════════════════════════════════════════════════════════════════

def detect_peaks(rows, column, threshold, min_gap):
    """Find peak indices where `column` exceeds threshold, with min_gap
    between peaks. Within each run of above-threshold frames, keep the
    frame with the highest value."""
    above = [(i, r[column]) for i, r in enumerate(rows) if r[column] >= threshold]
    if not above:
        return []

    peaks = []
    peak_idx, peak_val = above[0]
    prev_t = rows[above[0][0]]["timestamp"]

    for idx, val in above[1:]:
        t = rows[idx]["timestamp"]
        if t - prev_t > min_gap:
            peaks.append(peak_idx)
            peak_idx, peak_val = idx, val
        elif val > peak_val:
            peak_idx, peak_val = idx, val
        prev_t = t

    peaks.append(peak_idx)
    return peaks


def find_events(rows, startle_thr, fear_thr, min_gap, near_radius=3.0):
    """Two-pass event detection:
    Pass 1: MP startle peaks (primary, sharp physical reactions)
    Pass 2: HS fear peaks not near any startle peak (supplementary)
    Returns list of (peak_index, source_label) sorted by timestamp."""

    startle_peaks = detect_peaks(rows, "mp_startle_score", startle_thr, min_gap)
    hs_peaks = detect_peaks(rows, "hs_fear", fear_thr, min_gap)
    startle_times = {rows[p]["timestamp"] for p in startle_peaks}

    hs_only = []
    for p in hs_peaks:
        t = rows[p]["timestamp"]
        if not any(abs(t - st) < near_radius for st in startle_times):
            hs_only.append(p)

    events = [(p, "MP_STARTLE") for p in startle_peaks]
    events += [(p, "HS_FEAR") for p in hs_only]
    events.sort(key=lambda x: rows[x[0]]["timestamp"])
    return events


def get_window_indices(rows, center_t, radius):
    """Return (start_idx, end_idx) for frames within center_t +/- radius."""
    lo = center_t - radius
    hi = center_t + radius
    s_idx = None
    e_idx = None
    for i, r in enumerate(rows):
        if r["timestamp"] >= lo and s_idx is None:
            s_idx = i
        if r["timestamp"] <= hi:
            e_idx = i
    if s_idx is None:
        s_idx = 0
    if e_idx is None:
        e_idx = len(rows) - 1
    return s_idx, e_idx


def extract_event_v1(rows, s_idx, e_idx, peak_idx, event_id, source):
    """Extract stats for a single event window (v1: fixed window)."""
    cluster = rows[s_idx: e_idx + 1]
    n_total = len(cluster)

    cluster_start = cluster[0]["timestamp"]
    cluster_end = cluster[-1]["timestamp"]
    duration = cluster_end - cluster_start
    peak_t = rows[peak_idx]["timestamp"]

    # MP signals
    peak_startle = max(r["mp_startle_score"] for r in cluster)
    peak_tension = max(r["mp_tension"] for r in cluster)
    ctx_tags = sorted({r["mp_ctx_tag"] for r in cluster if r["mp_ctx_tag"] != "---"})

    # HS signals
    peak_hs_fear = max(r["hs_fear"] for r in cluster)
    peak_hs_surprise = max(r["hs_surprise"] for r in cluster)
    peak_arousal = max(r["hs_arousal"] for r in cluster)
    hs_dom_at_peak = rows[peak_idx]["hs_dominant"]

    # Per-emotion frame counts
    emotion_counts = _count_emotions(cluster)

    # Fusion
    peak_composite = max(r["composite_fear"] for r in cluster)
    n_veto = sum(1 for r in cluster if r["agreement_tag"] == "VETO")
    veto_set = sorted({r["veto_tag"] for r in cluster if r["veto_tag"] != "---"})

    # Detection gaps
    n_no_face = sum(1 for r in cluster
                    if r["mp_face_detected"] == 0 and r["hs_face_detected"] == 0)
    no_face_pct = (n_no_face / n_total * 100) if n_total > 0 else 0.0

    # Auto-classify cluster emotion
    n_fear = emotion_counts["n_fear_frames"]
    if n_fear > 0:
        cluster_emotion = "Fear"
    elif emotion_counts["n_surprise_frames"] > 0:
        cluster_emotion = "Surprise"
    else:
        cluster_emotion = "Startle"

    row = {
        "event_id":             event_id,
        "cluster_start":        f"{cluster_start:.3f}",
        "cluster_end":          f"{cluster_end:.3f}",
        "peak_t":               f"{peak_t:.3f}",
        "duration_s":           f"{duration:.2f}",
        "cluster_emotion":      cluster_emotion,
        "source":               source,
        # MP
        "peak_startle":         f"{peak_startle:.3f}",
        "peak_tension":         f"{peak_tension:.4f}",
        "mp_ctx_tags":          ",".join(ctx_tags) if ctx_tags else "---",
        # HS
        "peak_hs_fear":         f"{peak_hs_fear:.4f}",
        "peak_hs_surprise":     f"{peak_hs_surprise:.4f}",
        "peak_arousal":         f"{peak_arousal:.4f}",
        "hs_dominant_at_peak":  hs_dom_at_peak,
        # Fusion
        "peak_composite":       f"{peak_composite:.4f}",
        "n_veto_frames":        str(n_veto),
        "veto_reasons":         ",".join(veto_set) if veto_set else "---",
        # Gaps
        "no_face_pct":          f"{no_face_pct:.1f}",
        "n_no_face_frames":     str(n_no_face),
        # Manual (pre-fill start_s/end_s with cluster bounds)
        "verdict":              "",
        "category":             "",
        "start_s":              f"{cluster_start:.1f}",
        "end_s":                f"{cluster_end:.1f}",
        "context":              "",
        "face_visible_in_gaps": "",
    }
    # Add all emotion frame counts
    for col, count in emotion_counts.items():
        row[col] = str(count)

    return row


def run_v1(args, rows, csv_path):
    """Run v1 clustering: peak detection + fixed window."""
    print("  Mode: v1 (peak detection + fixed 6s window)")

    events_raw = find_events(rows, args.startle_threshold, args.fear_threshold,
                             args.min_gap)
    print(f"  Peaks found: {len(events_raw)} "
          f"(startle >= {args.startle_threshold}, hs_fear >= {args.fear_threshold}, "
          f"min_gap={args.min_gap}s)")

    n_startle = sum(1 for _, s in events_raw if s == "MP_STARTLE")
    n_hs = sum(1 for _, s in events_raw if s == "HS_FEAR")
    print(f"    MP startle peaks: {n_startle}  |  HS-only fear peaks: {n_hs}")

    results = []
    for i, (peak_idx, source) in enumerate(events_raw):
        center_t = rows[peak_idx]["timestamp"]
        s_idx, e_idx = get_window_indices(rows, center_t, args.window)
        eid = f"E{i + 1:02d}"
        results.append(extract_event_v1(rows, s_idx, e_idx, peak_idx, eid, source))

    print(f"  Events: {len(results)} (window={args.window}s radius)")

    # Summary
    n_fear = sum(1 for ev in results if ev["cluster_emotion"] == "Fear")
    n_surp = sum(1 for ev in results if ev["cluster_emotion"] == "Surprise")
    n_star = sum(1 for ev in results if ev["cluster_emotion"] == "Startle")
    n_veto_ev = sum(1 for ev in results if int(ev["n_veto_frames"]) > 0)
    print(f"\n  Cluster emotions: {n_fear} Fear, {n_surp} Surprise, {n_star} Startle")
    print(f"  Events with VETO: {n_veto_ev}")

    durations = [float(ev["duration_s"]) for ev in results]
    if durations:
        avg_dur = sum(durations) / len(durations)
        max_dur = max(durations)
        print(f"  Avg duration: {avg_dur:.1f}s  |  Max: {max_dur:.1f}s")

    results.extend(_make_fn_rows(HEADER_V1, args.fn_rows))

    out_dir = _output_dir(csv_path)
    out_path = out_dir / f"{csv_path.stem}_v1_annotation.csv"
    _write_csv(results, HEADER_V1, out_path)

    n_events = len(results) - args.fn_rows
    print(f"\n  Output: {out_path}")
    print(f"  Total rows: {len(results)} ({n_events} events + {args.fn_rows} FN)")

    print("\nNext steps:")
    print("  1. Open the template CSV in a spreadsheet editor")
    print("  2. Open the video file in VLC (Ctrl+T to jump to timestamps)")
    print("  3. For each event: watch peak_t +/- 2s, fill verdict/category/context")
    print("  4. Quick full-pass at 2x speed for missed events -> add in FN rows")
    print(f"  5. Save completed file as: {csv_path.stem}_v1_gt.csv")

    return out_path


# ══════════════════════════════════════════════════════════════════════════
# v2: Flood-fill clustering with confidence scoring
# ══════════════════════════════════════════════════════════════════════════

def flood_fill_clusters(rows, gap_tolerance, min_frames,
                        seed_mode="dominant", fear_thr=0.15, surprise_thr=0.25):
    """Chain consecutive seed frames within gap_tolerance into
    adaptive-width clusters.

    seed_mode:
      "dominant"  -- seed = hs_dominant in {Fear, Surprise} (original)
      "threshold" -- seed = hs_fear >= fear_thr OR hs_surprise >= surprise_thr
                     (catches sub-dominant fear, produces more clusters)

    Algorithm:
      1. Collect all seed frame indices
      2. Walk sorted list; split at gaps > gap_tolerance
      3. For each cluster, find peak (highest composite_fear)
      4. Drop clusters with fewer than min_frames seed frames

    Returns list of (s_idx, e_idx, peak_idx) sorted by timestamp.
    """
    if seed_mode == "threshold":
        seed_indices = sorted(
            i for i, r in enumerate(rows)
            if r["hs_fear"] >= fear_thr or r["hs_surprise"] >= surprise_thr
        )
    else:
        seed_indices = sorted(
            i for i, r in enumerate(rows) if r["hs_dominant"] in SEED_EMOTIONS
        )

    if not seed_indices:
        return []

    runs = []
    current_run = [seed_indices[0]]

    for idx in seed_indices[1:]:
        gap = rows[idx]["timestamp"] - rows[current_run[-1]]["timestamp"]
        if gap <= gap_tolerance:
            current_run.append(idx)
        else:
            runs.append(current_run)
            current_run = [idx]
    runs.append(current_run)

    clusters = []
    for indices in runs:
        if len(indices) >= min_frames:
            peak_idx = max(indices, key=lambda i: rows[i]["composite_fear"])
            clusters.append((indices[0], indices[-1], peak_idx))

    return clusters


def _padded_window(rows, s_idx, e_idx, pad_s=3.0):
    """Return (pad_s_idx, pad_e_idx) for cluster padded by +-pad_s seconds.
    Used for confidence scoring -- includes surrounding context frames so
    that short emotion bursts inside neutral periods score LOW."""
    t_lo = rows[s_idx]["timestamp"] - pad_s
    t_hi = rows[e_idx]["timestamp"] + pad_s
    pad_s_idx = s_idx
    pad_e_idx = e_idx
    for i in range(s_idx - 1, -1, -1):
        if rows[i]["timestamp"] >= t_lo:
            pad_s_idx = i
        else:
            break
    for i in range(e_idx + 1, len(rows)):
        if rows[i]["timestamp"] <= t_hi:
            pad_e_idx = i
        else:
            break
    return pad_s_idx, pad_e_idx


def extract_event_v2(rows, s_idx, e_idx, peak_idx, event_id,
                     sustained_thr, arousal_thr, sustained_pct_thr):
    """Extract stats for a single flood-fill cluster."""
    cluster = rows[s_idx: e_idx + 1]
    n_total = len(cluster)

    cluster_start = cluster[0]["timestamp"]
    cluster_end = cluster[-1]["timestamp"]
    duration = cluster_end - cluster_start
    peak_t = rows[peak_idx]["timestamp"]

    # MP signals
    peak_startle = max(r["mp_startle_score"] for r in cluster)
    peak_tension = max(r["mp_tension"] for r in cluster)
    ctx_tags = sorted({r["mp_ctx_tag"] for r in cluster if r["mp_ctx_tag"] != "---"})

    # HS signals
    peak_hs_fear = max(r["hs_fear"] for r in cluster)
    peak_hs_surprise = max(r["hs_surprise"] for r in cluster)
    peak_arousal = max(r["hs_arousal"] for r in cluster)
    hs_dom_at_peak = rows[peak_idx]["hs_dominant"]

    # Per-emotion frame counts
    emotion_counts = _count_emotions(cluster)

    # Fusion
    peak_composite = max(r["composite_fear"] for r in cluster)
    n_veto = sum(1 for r in cluster if r["agreement_tag"] == "VETO")
    veto_set = sorted({r["veto_tag"] for r in cluster if r["veto_tag"] != "---"})

    # Detection gaps
    n_no_face = sum(1 for r in cluster
                    if r["mp_face_detected"] == 0 and r["hs_face_detected"] == 0)
    no_face_pct = (n_no_face / n_total * 100) if n_total > 0 else 0.0

    # Confidence scoring -- use PADDED window (cluster +- 3s) so that
    # short emotion bursts inside neutral periods dilute to LOW.
    pad_s, pad_e = _padded_window(rows, s_idx, e_idx)
    pad_cluster = rows[pad_s: pad_e + 1]
    n_pad = len(pad_cluster)

    n_sustained = sum(1 for r in pad_cluster if r["composite_fear"] >= sustained_thr)
    sustained_pct = (n_sustained / n_pad * 100) if n_pad > 0 else 0.0
    mean_arousal = sum(r["hs_arousal"] for r in pad_cluster) / n_pad if n_pad > 0 else 0.0

    if sustained_pct >= sustained_pct_thr or mean_arousal >= arousal_thr:
        auto_confidence = "HIGH"
    else:
        auto_confidence = "LOW"

    # Auto-classify cluster emotion
    n_fear = emotion_counts["n_fear_frames"]
    if n_fear > 0:
        cluster_emotion = "Fear"
    elif emotion_counts["n_surprise_frames"] > 0:
        cluster_emotion = "Surprise"
    else:
        cluster_emotion = "Unknown"

    # Source: describe which signals contributed
    if peak_startle >= 2.5 and n_fear > 0:
        source = "MP+HS"
    elif peak_startle >= 2.5:
        source = "MP+HS_SURP"
    elif n_fear > 0:
        source = "HS_FEAR"
    else:
        source = "HS_SURP"

    row = {
        "event_id":             event_id,
        "cluster_start":        f"{cluster_start:.3f}",
        "cluster_end":          f"{cluster_end:.3f}",
        "peak_t":               f"{peak_t:.3f}",
        "duration_s":           f"{duration:.2f}",
        "cluster_emotion":      cluster_emotion,
        "source":               source,
        # Manual (annotator fills these)
        "label":                "",
        "start_s":              "",
        "end_s":                "",
        "context":              "",
        # Auto
        "auto_confidence":      auto_confidence,
        "face_visible_in_gaps": "",
        # Fusion + confidence
        "peak_composite":       f"{peak_composite:.4f}",
        "sustained_pct":        f"{sustained_pct:.1f}",
        "mean_arousal":         f"{mean_arousal:.4f}",
        "n_veto_frames":        str(n_veto),
        "veto_reasons":         ",".join(veto_set) if veto_set else "---",
        # HS
        "hs_dominant_at_peak":  hs_dom_at_peak,
        "peak_hs_fear":         f"{peak_hs_fear:.4f}",
        "peak_hs_surprise":     f"{peak_hs_surprise:.4f}",
        "peak_arousal":         f"{peak_arousal:.4f}",
        # MP
        "peak_startle":         f"{peak_startle:.3f}",
        "peak_tension":         f"{peak_tension:.4f}",
        "mp_ctx_tags":          ",".join(ctx_tags) if ctx_tags else "---",
        # Gaps
        "no_face_pct":          f"{no_face_pct:.1f}",
        "n_no_face_frames":     str(n_no_face),
    }
    # Add all emotion frame counts
    for col, count in emotion_counts.items():
        row[col] = str(count)

    return row


def run_v2(args, rows, csv_path):
    """Run v2 clustering: flood-fill with confidence scoring."""
    print(f"  Mode: v2 (flood-fill clustering, formula={args.formula})")

    # Count seed frames
    if args.seed == "threshold":
        n_seeds = sum(1 for r in rows
                      if r["hs_fear"] >= args.fear_seed_thr
                      or r["hs_surprise"] >= args.surprise_seed_thr)
        print(f"  Seed mode: threshold (hs_fear>={args.fear_seed_thr} "
              f"OR hs_surprise>={args.surprise_seed_thr})")
    else:
        n_seeds = sum(1 for r in rows if r["hs_dominant"] in SEED_EMOTIONS)
        print(f"  Seed mode: dominant (hs_dominant in Fear/Surprise)")
    print(f"  Seed frames: {n_seeds} / {len(rows)}")

    raw_clusters = flood_fill_clusters(
        rows, args.gap_tolerance, args.min_frames,
        seed_mode=args.seed,
        fear_thr=args.fear_seed_thr,
        surprise_thr=args.surprise_seed_thr,
    )
    print(f"  Clusters found: {len(raw_clusters)} "
          f"(gap_tolerance={args.gap_tolerance}s, min_frames={args.min_frames})")

    results = []
    for i, (s_idx, e_idx, peak_idx) in enumerate(raw_clusters):
        eid = f"E{i + 1:02d}"
        results.append(extract_event_v2(
            rows, s_idx, e_idx, peak_idx, eid,
            args.sustained_threshold,
            args.arousal_threshold,
            args.sustained_pct_threshold,
        ))

    # Summary
    n_fear_cl = sum(1 for ev in results if ev["cluster_emotion"] == "Fear")
    n_surp_cl = sum(1 for ev in results if ev["cluster_emotion"] == "Surprise")
    n_high = sum(1 for ev in results if ev["auto_confidence"] == "HIGH")
    n_low = sum(1 for ev in results if ev["auto_confidence"] == "LOW")
    n_veto_ev = sum(1 for ev in results if int(ev["n_veto_frames"]) > 0)
    print(f"\n  Cluster emotions: {n_fear_cl} Fear, {n_surp_cl} Surprise")
    print(f"  Confidence: {n_high} HIGH, {n_low} LOW")
    print(f"  Events with VETO: {n_veto_ev}")

    durations = [float(ev["duration_s"]) for ev in results]
    if durations:
        avg_dur = sum(durations) / len(durations)
        min_dur = min(durations)
        max_dur = max(durations)
        print(f"  Duration: avg={avg_dur:.1f}s  min={min_dur:.1f}s  max={max_dur:.1f}s")

    results.extend(_make_fn_rows(HEADER_V2, args.fn_rows))

    out_dir = _output_dir(csv_path)
    out_path = out_dir / f"{csv_path.stem}_v2_annotation.csv"
    _write_csv(results, HEADER_V2, out_path)

    n_events = len(results) - args.fn_rows
    print(f"\n  Output: {out_path}")
    print(f"  Total rows: {len(results)} ({n_events} events + {args.fn_rows} FN)")

    print("\nNext steps:")
    print("  1. Open the template CSV in a spreadsheet editor")
    print("  2. Open the video file in VLC (Ctrl+T to jump to timestamps)")
    print("  3. For each event: watch peak_t, fill label/context")
    print("     - label: Fear / Surprise / Neutral / Angry / ...")
    print("     - start_s / end_s: manual timing (when the emotion begins/ends)")
    print("     HIGH confidence -> quick confirm/dismiss")
    print("     LOW confidence  -> likely FP, verify quickly")
    print("  4. Quick full-pass at 2x speed for missed events -> add in FN rows")
    print(f"  5. Save completed file as: {csv_path.stem}_v2_gt.csv")

    return out_path


# ══════════════════════════════════════════════════════════════════════════
# v3: Multi-channel voting with independent signal channels
# ══════════════════════════════════════════════════════════════════════════

_TWO_GATE_ACTIVE = frozenset(("ONSET", "SUSTAINING", "EVENT_CONFIRMED"))


def build_channels(fear_thr=0.15, tension_thr=0.20, startle_thr=2.0,
                   arousal_thr=0.30, cross_fear_thr=0.10, cross_tension_thr=0.10):
    """Build 8 independent detection channels with configurable thresholds.

    Each channel uses a different signal modality so that votes are genuinely
    independent.  A HS-only false positive (face misread) fires CH1/CH3/CH6
    but not the MP channels; a MP-only false positive fires CH4/CH5 only.
    TWO_GATE fires only when the two-gate detector confirmed an onset event.
    """
    return [
        ("DOM_FEAR",      lambda r: r["hs_dominant"] == "Fear"),
        ("DOM_SURP",      lambda r: r["hs_dominant"] == "Surprise"),
        ("FEAR_THR",      lambda r, t=fear_thr: r["hs_fear"] > t),
        ("TENSION",       lambda r, t=tension_thr: r["mp_tension"] > t),
        ("STARTLE",       lambda r, t=startle_thr: r["mp_startle_score"] > t),
        ("AROUSAL",       lambda r, t=arousal_thr: r["hs_arousal"] > t),
        ("CROSS_MODAL",   lambda r, ft=cross_fear_thr, tt=cross_tension_thr:
                              r["hs_fear"] > ft and r["mp_tension"] > tt),
        ("TWO_GATE",      lambda r: r.get("event_status", "IDLE") in _TWO_GATE_ACTIVE),
    ]


def detect_scene_cuts(video_path, threshold=0.4, sample_step=1):
    """Detect scene cuts via grayscale histogram Bhattacharyya distance.

    Returns list of cut timestamps in seconds (sorted).
    """
    try:
        import cv2
    except ImportError:
        print("  WARNING: cv2 not available, skipping auto scene-cut detection")
        return []

    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        print(f"  WARNING: Could not open video: {video_path}")
        return []

    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    cuts = []
    prev_hist = None
    frame_idx = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        if frame_idx % sample_step == 0:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            hist = cv2.calcHist([gray], [0], None, [64], [0, 256])
            cv2.normalize(hist, hist)

            if prev_hist is not None:
                diff = cv2.compareHist(prev_hist, hist, cv2.HISTCMP_BHATTACHARYYA)
                if diff > threshold:
                    cuts.append(frame_idx / fps)

            prev_hist = hist

        frame_idx += 1

    cap.release()
    print(f"  Scene cuts auto-detected: {len(cuts)} (threshold={threshold})")
    return cuts


def flood_fill_channel(rows, seed_fn, gap_tolerance=0.5, min_frames=2,
                       scene_cuts=None):
    """Flood-fill clustering for a single detection channel.

    Returns list of (start_idx, end_idx) tuples representing intervals
    where this channel is active.
    """
    seed_indices = [i for i, r in enumerate(rows) if seed_fn(r)]
    if not seed_indices:
        return []

    # Chain into runs, splitting at gaps and scene cuts
    runs = []
    current_run = [seed_indices[0]]

    for idx in seed_indices[1:]:
        prev_t = rows[current_run[-1]]["timestamp"]
        curr_t = rows[idx]["timestamp"]
        gap = curr_t - prev_t

        # Check if a scene cut falls between prev and current
        cut_between = False
        if scene_cuts:
            for cut_t in scene_cuts:
                if prev_t < cut_t <= curr_t:
                    cut_between = True
                    break

        if gap > gap_tolerance or cut_between:
            runs.append(current_run)
            current_run = [idx]
        else:
            current_run.append(idx)
    runs.append(current_run)

    intervals = []
    for indices in runs:
        if len(indices) >= min_frames:
            intervals.append((indices[0], indices[-1]))

    return intervals


def compute_vote_counts(rows, channel_results):
    """Compute per-frame vote counts from all channel interval lists.

    channel_results: list of (channel_name, [(start_idx, end_idx), ...])
    Returns list of ints, length = len(rows).
    """
    n = len(rows)
    votes = [0] * n
    for _name, intervals in channel_results:
        for (s_idx, e_idx) in intervals:
            for i in range(s_idx, min(e_idx + 1, n)):
                votes[i] += 1
    return votes


def consolidate_vote_clusters(rows, votes, vote_candidate, vote_confirm,
                              gap_tolerance, min_duration, fps):
    """Group consecutive frames with vote_count >= vote_candidate into
    clusters.  Merge across small gaps.  Assign tier by peak vote count.

    Returns list of cluster dicts.
    """
    n = len(rows)
    gap_frames = int(gap_tolerance * fps)

    # Find runs of frames with votes >= vote_candidate
    runs = []
    current_run = None
    for i in range(n):
        if votes[i] >= vote_candidate:
            if current_run is None:
                current_run = [i]
            else:
                current_run.append(i)
        else:
            if current_run is not None:
                runs.append(current_run)
                current_run = None
    if current_run is not None:
        runs.append(current_run)

    if not runs:
        return []

    # Merge runs separated by small gaps
    merged = [runs[0]]
    for run in runs[1:]:
        prev_end = merged[-1][-1]
        curr_start = run[0]
        if curr_start - prev_end <= gap_frames:
            merged[-1].extend(run)
        else:
            merged.append(run)

    clusters = []
    for run in merged:
        s_idx = run[0]
        e_idx = run[-1]

        duration = rows[e_idx]["timestamp"] - rows[s_idx]["timestamp"]
        if duration < min_duration:
            continue

        peak_idx = max(range(s_idx, e_idx + 1), key=lambda i: votes[i])
        peak_count = votes[peak_idx]
        span_len = e_idx - s_idx + 1
        mean_count = sum(votes[i] for i in range(s_idx, e_idx + 1)) / span_len

        tier = "CONFIRMED" if peak_count >= vote_confirm else "CANDIDATE"

        clusters.append({
            "s_idx": s_idx,
            "e_idx": e_idx,
            "peak_vote_idx": peak_idx,
            "tier": tier,
            "peak_vote_count": peak_count,
            "mean_vote_count": mean_count,
        })

    return clusters


def extract_event_v3(rows, cluster, channels, scene_cuts, event_id,
                     channel_results=None, rppg_stats=None,
                     formula_vote_thr=FORMULA_VOTE_THR):
    """Extract stats for a single v3 voting cluster."""
    s_idx = cluster["s_idx"]
    e_idx = cluster["e_idx"]
    peak_idx = cluster["peak_vote_idx"]

    span = rows[s_idx:e_idx + 1]
    n_total = len(span)

    start_t = rows[s_idx]["timestamp"]
    end_t = rows[e_idx]["timestamp"]
    peak_t = rows[peak_idx]["timestamp"]
    duration = end_t - start_t

    # All benchmark formula scores across the full cluster span
    formula_scores = _all_formula_scores(span)

    # Gate overlap: what fraction of cluster frames the two-gate detector confirmed
    _gate_active = frozenset(("ONSET", "SUSTAINING", "EVENT_CONFIRMED"))
    gate_active_count = sum(1 for r in span if r.get("event_status", "IDLE") in _gate_active)
    gate_overlap_pct = (gate_active_count / n_total * 100) if n_total > 0 else 0.0

    # Channels whose intervals contain the peak frame (matches vote_count)
    if channel_results:
        active = []
        for name, intervals in channel_results:
            for (iv_s, iv_e) in intervals:
                if iv_s <= peak_idx <= iv_e:
                    active.append(name)
                    break
    else:
        active = _channels_active_at(rows[peak_idx], channels)

    # HS signals at peak
    peak_row = rows[peak_idx]
    cluster_emotion = peak_row["hs_dominant"]

    # Per-emotion counts
    emotion_counts = _count_emotions(span)

    # Data quality
    n_no_face = sum(1 for r in span
                    if r["mp_face_detected"] == 0 and r["hs_face_detected"] == 0)
    no_face_pct = (n_no_face / n_total * 100) if n_total > 0 else 0.0
    n_veto = sum(1 for r in span if r["agreement_tag"] == "VETO")

    # Scene cut bounded: cluster start or end within 0.1s of a cut
    bounded = False
    if scene_cuts:
        for cut_t in scene_cuts:
            if abs(start_t - cut_t) < 0.1 or abs(end_t - cut_t) < 0.1:
                bounded = True
                break

    # channels_active with HS/MP/BOTH prefix for source clarity
    labeled_active = [
        f"{CHANNEL_SOURCE.get(n, '?')}:{n}" for n in active
    ]

    # channels_filtered: active channels whose emotion vote matches cluster_emotion
    filtered = [n for n in active if CHANNEL_EMOTION_VOTE.get(n) == cluster_emotion]
    labeled_filtered = [f"{CHANNEL_SOURCE.get(n, '?')}:{n}" for n in filtered]

    # formulas_voted: formula names with peak score >= threshold
    voted_names = [
        f"F{i}" for i in range(7)
        if formula_scores.get(f"peak_f{i}", 0) >= formula_vote_thr
    ]

    row = {
        "event_id":           event_id,
        "tier":               cluster["tier"],
        "cluster_start_s":    f"{start_t:.3f}",
        "cluster_end_s":      f"{end_t:.3f}",
        "duration_s":         f"{duration:.2f}",
        "peak_t":             f"{peak_t:.3f}",
        # System emotion assessment
        "cluster_emotion":    cluster_emotion,
        # Manual annotation (blank for annotator)
        "verdict":            "",
        "category":           "",
        "context":            "",
        "rppg_watch":         "",
        "start_s":            "",
        "end_s":              "",
        # Per-formula emotion vote: cluster_emotion if formula fired, "---" if not
        **{f"F{i}_vote": (cluster_emotion
                          if formula_scores.get(f"peak_f{i}", 0) >= formula_vote_thr
                          else "---")
           for i in range(7)},
        # Formula and channel summaries
        "formulas_voted":     ",".join(voted_names) if voted_names else "---",
        "channels_filtered":  ",".join(labeled_filtered) if labeled_filtered else "---",
        # Voting summary
        "vote_count_peak":    str(cluster["peak_vote_count"]),
        "vote_count_mean":    f"{cluster['mean_vote_count']:.1f}",
        "channels_active":    ",".join(labeled_active) if labeled_active else "---",
        # Per-channel binary (0/1)
        "ch_DOM_FEAR":        "1" if "DOM_FEAR"    in active else "0",
        "ch_DOM_SURP":        "1" if "DOM_SURP"    in active else "0",
        "ch_FEAR_THR":        "1" if "FEAR_THR"    in active else "0",
        "ch_TENSION":         "1" if "TENSION"     in active else "0",
        "ch_STARTLE":         "1" if "STARTLE"      in active else "0",
        "ch_AROUSAL":         "1" if "AROUSAL"      in active else "0",
        "ch_CROSS_MODAL":     "1" if "CROSS_MODAL"  in active else "0",
        "ch_TWO_GATE":        "1" if "TWO_GATE"     in active else "0",
        "gate_overlap_pct":   f"{gate_overlap_pct:.1f}",
        # Raw formula scores
        **{k: f"{v:.4f}" for k, v in formula_scores.items()},
        # HS signals at peak
        "peak_hs_fear":       f"{peak_row['hs_fear']:.4f}",
        "peak_hs_surprise":   f"{peak_row['hs_surprise']:.4f}",
        "peak_hs_arousal":    f"{peak_row['hs_arousal']:.4f}",
        # MP signals at peak
        "peak_mp_tension":    f"{peak_row['mp_tension']:.4f}",
        "peak_mp_startle":    f"{peak_row['mp_startle_score']:.3f}",
        # Data quality
        "no_face_pct":        f"{no_face_pct:.1f}",
        "n_veto_frames":      str(n_veto),
        "scene_cut_bounded":  str(bounded),
        # rPPG fields (from sidecar or blank)
        **(rppg_stats or {}),
        "rppg_impression":    "",
        "rppg_notes":         "",
    }
    for col, count in emotion_counts.items():
        row[col] = str(count)

    return row


def run_v3(args, rows, csv_path):
    """Run v3 clustering: multi-channel voting with 7 independent channels."""
    print("  Mode: v3 (multi-channel voting, 8 independent channels)")

    channels = build_channels(
        fear_thr=args.ch_fear_thr,
        tension_thr=args.ch_tension_thr,
        startle_thr=args.ch_startle_thr,
        arousal_thr=args.ch_arousal_thr,
        cross_fear_thr=args.ch_cross_fear_thr,
        cross_tension_thr=args.ch_cross_tension_thr,
    )

    # Determine mode
    mode = args.mode
    if mode == "auto":
        mode = "continuous"
        print("  Mode auto -> defaulting to continuous (use --mode compilation for scene cuts)")

    # rPPG sidecar loading (optional)
    rppg_data = {a: [] for a in RPPG_ALGOS}
    rppg_csv = getattr(args, "rppg_csv", None)
    if not rppg_csv:
        # Auto-detect sidecar: same stem + _rppg.csv
        sidecar_rppg = csv_path.with_name(csv_path.stem + "_rppg.csv")
        if sidecar_rppg.exists():
            rppg_csv = str(sidecar_rppg)
            print(f"  rPPG sidecar auto-detected: {sidecar_rppg.name}")
    if rppg_csv:
        rppg_data = load_rppg_csv(rppg_csv)
        for algo in RPPG_ALGOS:
            n = len(rppg_data[algo])
            print(f"    {algo}: {n} BPM windows loaded")
    else:
        print("  rPPG: no sidecar found — rppg_* columns will be blank (fill manually)")

    # Scene cut detection (compilation mode only)
    scene_cuts = []
    if mode == "compilation":
        auto_cuts = []
        if args.video:
            auto_cuts = detect_scene_cuts(args.video, threshold=0.4)
        else:
            print("  WARNING: --video not provided, no auto scene-cut detection")

        cuts_file = args.cuts_file
        # Check for sidecar <csv_stem>.cuts.txt
        sidecar = csv_path.with_suffix(".cuts.txt")
        if not cuts_file and sidecar.exists():
            cuts_file = str(sidecar)

        scene_cuts = load_scene_cuts(cuts_file, auto_cuts)
        if scene_cuts:
            print(f"  Scene cuts (merged): {len(scene_cuts)} boundaries")

    # Estimate FPS from timestamp spacing
    if len(rows) >= 2:
        total_time = rows[-1]["timestamp"] - rows[0]["timestamp"]
        fps = len(rows) / total_time if total_time > 0 else 30.0
    else:
        fps = 30.0

    # Per-channel flood-fill
    print(f"\n  Channel flood-fill (gap={args.gap_tolerance}s, min_frames={args.min_frames}):")
    channel_results = []
    for name, seed_fn in channels:
        intervals = flood_fill_channel(
            rows, seed_fn,
            gap_tolerance=args.gap_tolerance,
            min_frames=args.min_frames,
            scene_cuts=scene_cuts,
        )
        channel_results.append((name, intervals))
        n_intervals = len(intervals)
        n_frames = sum(e - s + 1 for s, e in intervals)
        print(f"    {name:18s}: {n_intervals:3d} intervals, {n_frames:5d} frames")

    # Temporal voting
    votes = compute_vote_counts(rows, channel_results)
    max_vote = max(votes) if votes else 0
    n_above_cand = sum(1 for v in votes if v >= args.vote_candidate)
    n_above_conf = sum(1 for v in votes if v >= args.vote_confirm)
    print(f"\n  Voting: max={max_vote}/8, "
          f"frames>={args.vote_candidate} (candidate): {n_above_cand}, "
          f"frames>={args.vote_confirm} (confirm): {n_above_conf}")

    # Consolidate into clusters
    clusters = consolidate_vote_clusters(
        rows, votes,
        vote_candidate=args.vote_candidate,
        vote_confirm=args.vote_confirm,
        gap_tolerance=args.gap_tolerance,
        min_duration=args.min_duration,
        fps=fps,
    )

    n_confirmed = sum(1 for c in clusters if c["tier"] == "CONFIRMED")
    n_candidate = sum(1 for c in clusters if c["tier"] == "CANDIDATE")
    print(f"  Clusters: {len(clusters)} total "
          f"({n_confirmed} CONFIRMED, {n_candidate} CANDIDATE)")

    # Extract events
    results = []
    for i, cluster in enumerate(clusters):
        if args.hide_candidate and cluster["tier"] != "CONFIRMED":
            continue
        eid = f"E{i + 1:02d}"
        start_s = rows[cluster["s_idx"]]["timestamp"]
        end_s   = rows[cluster["e_idx"]]["timestamp"]
        rppg_stats = rppg_cluster_stats(rppg_data, start_s, end_s)
        results.append(extract_event_v3(rows, cluster, channels, scene_cuts, eid,
                                        channel_results=channel_results,
                                        rppg_stats=rppg_stats,
                                        formula_vote_thr=args.formula_vote_thr))

    # Summary
    if results:
        durations = [float(ev["duration_s"]) for ev in results]
        avg_dur = sum(durations) / len(durations)
        print(f"  Duration: avg={avg_dur:.1f}s  "
              f"min={min(durations):.1f}s  max={max(durations):.1f}s")

        n_fear_cl = sum(1 for ev in results if ev["cluster_emotion"] == "Fear")
        n_surp_cl = sum(1 for ev in results if ev["cluster_emotion"] == "Surprise")
        n_veto_ev = sum(1 for ev in results if int(ev["n_veto_frames"]) > 0)
        print(f"  Emotions: {n_fear_cl} Fear, {n_surp_cl} Surprise")
        print(f"  Events with VETO: {n_veto_ev}")

    results.extend(_make_fn_rows(HEADER_V3, args.fn_rows))

    # Write output as .xlsx (freeze panes + category dropdown)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        _HAS_OPENPYXL = True
    except ImportError:
        _HAS_OPENPYXL = False

    out_dir = _output_dir(csv_path)

    if _HAS_OPENPYXL:
        out_path = out_dir / f"{csv_path.stem}_v3_annotation.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Annotation"

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        for ci, col in enumerate(HEADER_V3, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill

        for ri, row_data in enumerate(results, 2):
            for ci, col in enumerate(HEADER_V3, 1):
                ws.cell(row=ri, column=ci, value=row_data.get(col, ""))

        # Freeze row 1 (headers) and column A (event_id)
        ws.freeze_panes = "B2"

        # Category dropdown validation
        cat_col = get_column_letter(HEADER_V3.index("category") + 1)
        dv = DataValidation(
            type="list",
            formula1=f'"{CATEGORY_OPTIONS}"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{cat_col}2:{cat_col}{len(results) + 1}"

        # Auto column width (capped at 30 chars)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        wb.save(out_path)
    else:
        # Fallback to CSV if openpyxl is not installed
        out_path = out_dir / f"{csv_path.stem}_v3_annotation.csv"
        _write_csv(results, HEADER_V3, out_path)
        print("  WARNING: openpyxl not found — wrote CSV fallback (no freeze panes).")
        print("  Install with: pip install openpyxl")

    n_events = len(results) - args.fn_rows
    print(f"\n  Output: {out_path}")
    print(f"  Total rows: {len(results)} ({n_events} events + {args.fn_rows} FN)")

    print("\nNext steps:")
    print("  1. Open the .xlsx in Excel (row 1 + column A are frozen)")
    print("  2. Open the video in VLC (Ctrl+T to jump to timestamps in seconds)")
    print("  3. For each E-row: seek to peak_t, watch +-3s context")
    print("     CONFIRMED (vote_count_peak >=5) -> quick verify")
    print("     CANDIDATE (3-4) -> careful review")
    print("  4. Fill: verdict (observed emotion / FP / SKIP)")
    print("     then category (dropdown), rppg_watch, start_s/end_s, context")
    print("  5. Quick 2x speed full-pass for missed events -> fill FN rows")
    print(f"  6. Save completed file as: {csv_path.stem}_v3_gt.xlsx")

    return out_path
