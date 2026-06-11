"""One-off patch: update the category dropdown in an existing V3 annotation xlsx.
Usage:
    python patch_category_dropdown.py --xlsx path/to/_v3_annotation.xlsx
The file is updated in-place. All existing cell data is preserved.
"""
import argparse
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from annotate_helper import HEADER_V3, CATEGORY_OPTIONS


def patch(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Remove any existing data validations on the category column
    cat_col_idx = HEADER_V3.index("category") + 1
    cat_col = get_column_letter(cat_col_idx)
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if cat_col not in str(dv.sqref)
    ]

    # Add updated dropdown
    n_rows = ws.max_row
    dv = DataValidation(
        type="list",
        formula1=f'"{CATEGORY_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{cat_col}2:{cat_col}{n_rows + 1}"

    wb.save(xlsx_path)
    print(f"Updated category dropdown in: {xlsx_path}")
    print(f"Options: {CATEGORY_OPTIONS}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to annotation xlsx to patch")
    args = parser.parse_args()
    path = Path(args.xlsx).resolve()
    if not path.exists():
        print(f"ERROR: File not found: {path}")
        return
    patch(path)


if __name__ == "__main__":
    main()
